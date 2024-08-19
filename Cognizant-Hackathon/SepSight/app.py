from flask import Flask, render_template, request, redirect, url_for, flash
import sqlite3
from preprocessing import impute_logic
from model import create_model
import torch
import torch.nn.functional as F
import pandas as pd
from openpyxl import load_workbook  ,Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

"""Database: {'patient ID','Name','hour','HR', 'O2Sat','Temp','MAP','Resp','BUN','Chloride','Creatinine','Glucose','Hct','Hgb','WBC','Platelets','Age','Gender'"""
"""EXPECTED: {'HR', 'O2Sat','Temp','MAP','Resp','BUN','Chloride','Creatinine','Glucose','Hct','Hgb','WBC','Platelets','Age','Gender'"""


app = Flask(__name__)
app.secret_key = "super secret key"

def makeSuaitableFormat(data):
    if data=='':
        return data 
    else:
        return float(data)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/addPatientData', methods=['POST'])
def addPatientData():
    if request.method == 'POST':
        patient_id = request.form['patient_id']
        name = request.form['name']
        hour = makeSuaitableFormat(request.form['hour'])
        hr = makeSuaitableFormat(request.form['hr'])
        o2sat = makeSuaitableFormat(request.form['o2sat'])
        temp = makeSuaitableFormat(request.form['temp'])
        mp = makeSuaitableFormat(request.form['map'])
        resp = makeSuaitableFormat(request.form['resp'])
        bun = makeSuaitableFormat(request.form['bun'])
        chloride = makeSuaitableFormat(request.form['chloride'])
        creatinine = makeSuaitableFormat(request.form['creatinine'])
        glucose = makeSuaitableFormat(request.form['glucose'])
        hct = makeSuaitableFormat(request.form['hct'])
        hgb = makeSuaitableFormat(request.form['hgb'])
        wbc = makeSuaitableFormat(request.form['wbc'])
        platelets = makeSuaitableFormat(request.form['platelets'])
        age = makeSuaitableFormat(request.form['age'])
        gender = request.form['gender']
        print(patient_id,name,hour,hr,o2sat,temp,mp,resp,bun,chloride,creatinine,glucose,hct,hgb,wbc,platelets,age)
        conn = sqlite3.connect('sepsis.db', check_same_thread=False)
        cursor=conn.cursor()

        existdata = cursor.execute('''SELECT * FROM patients WHERE patient_id = ?''', (patient_id,)).fetchone()
        print(existdata)
        if existdata:
            cursor.execute(
                '''UPDATE patients SET hour=?, hr=?, o2sat=?, temp=?, mp=?, resp=?, bun=?, chloride=?, creatinine=?, glucose=?, hct=?, hgb=?, wbc=?, platelets=? 
                    WHERE patient_id=?''', 
                (
                    existdata[2]+','+str(hour), existdata[3]+','+str(hr), existdata[4]+','+str(o2sat), existdata[5]+','+str(temp), existdata[6]+','+str(mp), existdata[7]+','+str(resp), existdata[8]+','+str(bun), existdata[9]+','+str(chloride), existdata[10]+','+str(creatinine), existdata[11]+','+str(glucose), existdata[12]+','+str(hct), existdata[13]+','+str(hgb), existdata[14]+','+str(wbc), existdata[15]+','+str(platelets), patient_id
            )
            )
            
        else:
            if gender=="Male":
                gender='1'
            else:
                gender='0'
            cursor.execute(
        '''INSERT INTO patients (patient_id, name, hour, hr, o2sat, temp, mp, resp, bun, chloride, creatinine, glucose, hct, hgb, wbc, Platelets, age, gender) 
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', 
        (patient_id, name, hour, hr, o2sat, temp, mp, resp, bun, chloride, creatinine, glucose, hct, hgb, wbc, platelets, age, gender)
    )
        conn.commit()
        conn.close()

    flash("Patient data added successfully!")
    return redirect(url_for('index'))

def expandData(data):
    result=[]
    for record in data:
        timeStep=record[2].split(',')
        for t in range(len(timeStep)):
            temp=[record[0],record[1]]
            for x in range(2,16):
                splitData=record[x].split(',')
                temp.append(splitData[t])
            temp.append(record[16])
            temp.append(record[17])
            temp.append(t+1)
            result.append(temp)
    return result

@app.route('/predictSepsis/<patientid>/<name>/<timeStep>')
def predictSepsis(patientid,name,timeStep):
    print(patientid,name,timeStep)
    timeStep=int(timeStep)
    conn = sqlite3.connect('sepsis.db', check_same_thread=False)
    cursor=conn.cursor()
    data = cursor.execute('''SELECT * FROM patients WHERE patient_id = ?''',(patientid,)).fetchall()
    data=expandData(data)
    tdata=[]
    print(data)
    for i in range(timeStep):
        tdata.append(data[i])
    for i in tdata:
        for j in range(len(i)):
            if i[j]=='':
                i[j]='NaN'
            else:
                try:
                    i[j]=int(i[j])
                except:
                    pass
    modelData=[]
    print(tdata)
    for i in tdata:
        tp=[]
        for j in range(3,18):
            tp.append(i[j])
        modelData.append(tp)


    print(modelData)
    DataFrame = pd.DataFrame(modelData,columns=['HR','O2Sat','Temp','MAP','Resp','BUN','Chloride','Creatinine','Glucose','Hct','Hgb','WBC','Platelets','Age','Gender'])
    impute_logic(DataFrame)

    data_np = DataFrame.to_numpy()
    data_tensor = torch.from_numpy(data_np).long()
    print('Tensor Shape......')
    print(data_tensor.shape)
    model = create_model()

    model.eval()

    with torch.no_grad():
        y = model(data_tensor)

    # Apply softmax to get probabilities
    predictions = F.softmax(y, dim=-1)
    print(predictions)
    # Print probabilities
    probs = predictions.numpy()
    print('Predictions (Probabilities):')
    print(probs)
    print("++++++++++++++++++++++")
    print(probs)
    print(tdata)
    return render_template("sepsispredict.html",probs=probs,data=tdata) 



# def expandData(data):
#     result = []
#     for record in data:
#         id, name, *medical_values = record
        
#         # Determine the maximum length of the medical value lists
#         max_len = max(len(value.split(',')) for value in medical_values)
        
#         for i in range(max_len):
#             expanded_record = [id, name]
#             for value in medical_values:
#                 # Split the value and handle missing data by adding an empty string
#                 split_values = value.split(',')
#                 expanded_record.append(split_values[i] if i < len(split_values) else '')
#             expanded_record.append(i + 1)  # Add timestep
#             result.append(tuple(expanded_record))
    
#     return result

@app.route('/patientData')
def patientData():
    conn = sqlite3.connect('sepsis.db', check_same_thread=False)
    cursor=conn.cursor()
    data = cursor.execute('''SELECT * FROM patients''').fetchall()
    data=expandData(data)
    print(data)
    conn.close()
    return render_template('patientData.html', data=data)


@app.route('/predictCSV', methods=['POST'])
def predictCSV():
    if request.method == 'POST':
        file = request.files['file']
        #save
        file.save(file.filename)
        #read openpyxl
        wb = load_workbook(file.filename)
        sheet = wb.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        patients={}
        for idx in range(1, len(data)):
            data[idx] = list(data[idx])
            
            if data[idx][-1]=='Male':
                data[idx][-1]=1
            else:
                data[idx][-1]=0

            # Patient processing
            if data[idx][0] in patients:
                data[idx].append(len(patients[data[idx][0]])+1)
                patients[data[idx][0]].append(data[idx][3:18])
            else:
                data[idx].append(1)
                patients[data[idx][0]]=[data[idx][3:18]]   

        print("data:",data)
        probabilities=[]
        print(patients)
        for i in patients:
            currData=patients[i]
            print(currData)
            DataFrame = pd.DataFrame(currData,columns=['HR','O2Sat','Temp','MAP','Resp','BUN','Chloride','Creatinine','Glucose','Hct','Hgb','WBC','Platelets','Age','Gender'])
            impute_logic(DataFrame)

            data_np = DataFrame.to_numpy()
            data_tensor = torch.from_numpy(data_np).long()
            print('Tensor Shape......')
            print(data_tensor.shape)
            model = create_model()

            model.eval()

            with torch.no_grad():
                y = model(data_tensor)

            # Apply softmax to get probabilities
            predictions = F.softmax(y, dim=-1)
            print(predictions)
            # Print probabilities
            probs = predictions.numpy()
            print('Predictions (Probabilities):')
            for p in probs:
                probabilities.append(p)

        print(probabilities)
        print(patients)
        displayData=[]
        return render_template('sepsispredict.html', data=data[1:],probs=probabilities)

if __name__ == '__main__':
    app.run(debug=True,port=5000)
